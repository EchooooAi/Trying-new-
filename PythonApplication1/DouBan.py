import requests
from requests.exceptions import RequestException
from lxml import etree
import re 
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Fill

headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_4) '
                  'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36',}

def get_one_page(url):
    try:
        response = requests.get(url,headers = headers)
        return response
    except RequestException:
        return None

def parse_the_page(html):
    soup = BeautifulSoup(html,'lxml')

    # get everyone book's info
    #get_info = soup.find_all(name='dd')

    # get book's name and add to the list
    get_names = []
    for dd in soup.find_all(name='dd'):
        get_name = dd.find_all(name = 'a')
        for name in get_name:
             get_names.append(name.string)

    # get book's describe
    get_describes =[]
    for dd in soup.find_all(name = 'dd'):
        get_describes.append(dd.find_all(attrs = {'class':'desc'}))
    #正则 信息详细归类 暂时失败
    str_desc =''
    for describes in get_describes:
        for list_desc in describes:
            str_desc += list_desc.string
    #print(str_desc)
    # add the describe in list in order
    get_desc_list = (re.findall('\n.{8}(.*?)\n',str_desc,re.S))

    # using xpath to get the score
    #此处 最开始写成了 etree.parse(html,etree.HTMLParser()) 不正确 第一个形参应该传入的是一个本地html文件地址路径
    Xpath_html = etree.HTML(html)
    get_score = Xpath_html.xpath('//div[@class="rating"]//span[@class="rating_nums"]/text()')
    for i in range(15):
        yield {
            'name' : get_names[i] ,
            'describe' : get_desc_list[i] ,
            'score' : get_score[i],
            }

def write_to_xlsx(result):
    wb = Workbook()
    ws = wb.create_sheet(index=0,title='First Page')

    ws.freeze_panes = 'A2'
    col_A = ws.column_dimensions['A']
    col_A.width = 23
    # 这里无法应用于 每一个单元格，详情见文档 应该再单元格中逐一调整，如下面的ite_rows循环中
    #col_A.font = Font(name='方正舒体',bold=True)
    #row_1 = ws.row_dimensions[1]
    #row_1.alignment = Alignment(horizontal='center')

    list_title = ['Book\'s name','score', 'Details']
    for row in ws.iter_rows(min_row=1,max_row=1,max_col=3):
        i = 0
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
            cell.value =list_title[i]
            i = i + 1
    # from the first row , every row is a movie and initialize (name width = 22)
    list = ['name','score','describe']
    for row in ws.iter_rows(min_row=2,max_row =11,max_col=3):
        i = 0
        one_movie = result.__next__()
        for cell in row:
            cell.value = one_movie[list[i]]
            i = i + 1
    # 调整字体文本
    for col in ws.iter_cols(min_row=2,max_row=11,max_col=1):
        for cell in col:
            cell.font = Font(name='方正舒体',bold=True)

    return wb.save('sample.xlsx')
    
    

def main():
    url = 'https://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0'
    html = get_one_page(url).text
    results = parse_the_page(html)
    write_to_xlsx(results)

main()