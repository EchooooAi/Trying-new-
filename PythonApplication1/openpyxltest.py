from openpyxl import Workbook
from openpyxl.styles import colors, Color, Font, PatternFill, named_styles, NamedStyle, Border, Side

# https://openpyxl.readthedocs.io/en/latest/index.html

# 创建 xlsx实例
wb = Workbook()

# 创建默认 电子表格
ws = wb.active
ws_test = wb.create_sheet(index=1,title='test')
################
ws['A4'] = 4
ws['A1'] ='Hello'

########锁行 第一行 第一列
ws.freeze_panes = 'A2'
#ws.frezze_panes = 'B1'

########设定行列长宽
ws.column_dimensions['B'].width = 55
ws.row_dimensions[4].height = 70

#copy
source = wb.active
target = wb.copy_worksheet(source)

#  4行 2列 写入值 10
cel = ws.cell(row=4,column=2,value=10)

# 创建一行并 再cell中赋值
for row in ws.iter_rows(min_row=2,max_col=4,max_row=2):
    for cell in row:
        cell.value = 'Huiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii'

fill = PatternFill(fill_type=None,
                start_color='FFFFFFFF',
                end_color='FF000000')

# 名称风格
highlight = NamedStyle(name='hightlight')
highlight.font = Font(bold=True, size=20)
bd = Side(style='thick',color='000000')
highlight.border = Border(left=bd,top=bd,right=bd,bottom=bd)
wb.add_named_style(highlight)

ws_test['D5'] = 'hulllllll'
ws_test['E5'] = 'zzzzzzzz'
ws_test['D5'].style = highlight
########


wb.save('Fir.xlsx')